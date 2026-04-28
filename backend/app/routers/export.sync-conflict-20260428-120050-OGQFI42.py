import io
import os
import zipfile
from datetime import datetime

from fastapi import APIRouter, Depends, HTTPException
from fastapi.responses import StreamingResponse

from app.database import Database, DB_PATH
from app.dependencies import get_db

router = APIRouter(prefix="/api/export", tags=["export"])


@router.get("/db/sqlite")
def export_db_sqlite(db: Database = Depends(get_db)):
    if not os.path.exists(DB_PATH):
        raise HTTPException(status_code=404, detail="Banco de dados não encontrado")

    filename = f"gestao_contratos_{datetime.now().strftime('%Y%m%d_%H%M%S')}.db"

    def iterfile():
        with open(DB_PATH, "rb") as f:
            while chunk := f.read(8192):
                yield chunk

    return StreamingResponse(
        iterfile(),
        media_type="application/octet-stream",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'}
    )


@router.get("/db/xlsx")
def export_db_xlsx(db: Database = Depends(get_db)):
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
    except ImportError:
        raise HTTPException(status_code=500, detail="openpyxl não instalado")

    wb = Workbook()

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="4472C4")
    header_alignment = Alignment(horizontal="center", vertical="center")
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )

    def style_header(ws, row_num, num_cols):
        for col in range(1, num_cols + 1):
            cell = ws.cell(row=row_num, column=col)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = thin_border

    def add_data(ws, data, start_row):
        for row_idx, row in enumerate(data, start=start_row):
            for col_idx, value in enumerate(row, start=1):
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                cell.border = thin_border
                cell.alignment = Alignment(wrap_text=True)

    def auto_width(ws):
        for column in ws.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            for cell in column:
                try:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                except:
                    pass
            ws.column_dimensions[column_letter].width = min(max_length + 2, 50)

    clientes = db.get_all_clientes()
    if clientes:
        ws = wb.active
        ws.title = "Clientes"
        headers = ["ID", "Nome", "CPF/CNPJ", "Telefone", "Email", "CEP", "Logradouro",
                   "Número", "Complemento", "Cidade", "Estado", "Representante", "Observações"]
        ws.append(headers)
        style_header(ws, 1, len(headers))
        for c in clientes:
            ws.append([
                c["id"], c["nome"], c["cpf_cnpj"], c["telefone"], c["email"],
                c["cep"], c["logradouro"], c["numero"], c["complemento"],
                c["cidade"], c["estado"], c["nome_representante"], c["observacoes"]
            ])
        auto_width(ws)

    contratos_rows = db.conn.execute('''
        SELECT c.id, c.ctt_n, cl.nome AS cliente_nome, c.descricao, c.tipo,
               c.advogado, c.data_assinatura, c.status, c.observacoes, c.arquivo_path
        FROM contratos c JOIN clientes cl ON c.cliente_id = cl.id
        ORDER BY c.ctt_n
    ''').fetchall()

    if contratos_rows:
        ws = wb.create_sheet("Contratos")
        headers = ["ID", "CTT-N", "Cliente", "Descrição", "Tipo", "Advogado",
                   "Data Assinatura", "Status", "Observações", "Arquivo"]
        ws.append(headers)
        style_header(ws, 1, len(headers))
        for r in contratos_rows:
            ws.append([r["id"], r["ctt_n"], r["cliente_nome"], r["descricao"],
                       r["tipo"], r["advogado"], r["data_assinatura"], r["status"],
                       r["observacoes"], r["arquivo_path"]])
        auto_width(ws)

    honorarios_rows = db.conn.execute('''
        SELECT h.id, h.contrato_id, c.ctt_n, h.tipo, h.hipotese, h.valor
        FROM honorarios h JOIN contratos c ON h.contrato_id = c.id
        ORDER BY c.ctt_n, h.tipo, h.ordem
    ''').fetchall()

    if honorarios_rows:
        ws = wb.create_sheet("Honorários")
        headers = ["ID", "Contrato ID", "CTT-N", "Tipo", "Hipótese", "Valor"]
        ws.append(headers)
        style_header(ws, 1, len(headers))
        for r in honorarios_rows:
            ws.append([r["id"], r["contrato_id"], r["ctt_n"], r["tipo"],
                       r["hipotese"], r["valor"]])
        auto_width(ws)

    parcelas_rows = db.conn.execute('''
        SELECT p.id, p.honorario_id, h.tipo AS honorario_tipo, c.ctt_n,
               p.num_parcela, p.valor, p.vencimento, p.nota_fiscal, p.data_pagamento
        FROM parcelas p
        JOIN honorarios h ON p.honorario_id = h.id
        JOIN contratos c ON h.contrato_id = c.id
        ORDER BY c.ctt_n, h.tipo, p.num_parcela
    ''').fetchall()

    if parcelas_rows:
        ws = wb.create_sheet("Parcelas")
        headers = ["ID", "Honorário ID", "CTT-N", "Tipo Honorário",
                   "Nº Parcela", "Valor", "Vencimento", "Nota Fiscal", "Data Pagamento"]
        ws.append(headers)
        style_header(ws, 1, len(headers))
        for r in parcelas_rows:
            ws.append([r["id"], r["honorario_id"], r["ctt_n"], r["honorario_tipo"],
                       r["num_parcela"], r["valor"], r["vencimento"],
                       r["nota_fiscal"], r["data_pagamento"]])
        auto_width(ws)

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    filename = f"gestao_contratos_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"

    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'}
    )


@router.get("/contratos/zip")
def export_contratos_zip(db: Database = Depends(get_db)):
    contratos = db.conn.execute('''
        SELECT c.id, c.ctt_n, c.arquivo_path
        FROM contratos c
        ORDER BY c.ctt_n
    ''').fetchall()

    buffer = io.BytesIO()
    with zipfile.ZipFile(buffer, "w", zipfile.ZIP_DEFLATED) as zf:
        for contrato in contratos:
            if contrato["arquivo_path"] and os.path.exists(contrato["arquivo_path"]):
                filename = os.path.basename(contrato["arquivo_path"])
                safe_name = f"{contrato['ctt_n']}_{filename}"
                zf.write(contrato["arquivo_path"], safe_name)

    buffer.seek(0)
    filename = f"contratos_{datetime.now().strftime('%Y%m%d_%H%M%S')}.zip"

    return StreamingResponse(
        buffer,
        media_type="application/zip",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'}
    )